import openpyxl
import pandas as pd
import openpyxl as op
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import math
import numpy as np
import re
import webcolors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import worker

com_id = 0
total_id = 0


# Function to read the Excel file using pandas
def get_file(filepath):
    # Read the Excel file into a pandas DataFrame
    try:
        df = pd.DataFrame(pd.read_excel(filepath))
    except:
        print("Error, No such file founded: Arbeitspakete file")
        exit(1)
    return df


def get_workers_info(filepath, lista_months):
    worker.list_of_av_worker.clear()
    try:
        pf = pd.DataFrame(pd.read_excel(filepath))
    except:
        print("Error, No such file founded: Worker File")
        exit(1)
    list_workers = []
    for line in pf.values.__array__()[1:]:
        w = worker.Worker(line[0], line[2], len(lista_months))
        if type(line[1]) == str:
            number = line[1].split(",")
            num = float(number[0]) + float(number[1]) / 10
            w.allowed_hours(num, lista_months)
            list_workers.append(w)
            worker.list_of_av_worker.append(num)
            continue
        w.allowed_hours(float(line[1]), lista_months)
        list_workers.append(w)
        worker.list_of_av_worker.append(line[1])
    worker.add_to_list(list_workers)


# Function to extract numbers from the first column of the DataFrame
def get_nrs(df):
    numbers = []
    counter = 0
    for line in df.values.__array__():
        line_to_string = str(line[0])
        if line_to_string != 'nan':
            try:
                # Attempt to convert to float
                id_float = float(line_to_string)
                # Check if the number is an integer by comparing it to its int version
                if not id_float.is_integer():
                    if str(id_float) in numbers:
                        line_to_string = line_to_string + "0"
                        if len(line_to_string.split(".")[1]) > 2:
                            line_to_string = line_to_string[0:-1]
                    numbers.append(line_to_string)
            except ValueError:
                # If it cannot be converted to float, check if it is a hierarchical number
                if '.' in line_to_string:
                    numbers.append(line_to_string)
    return numbers


# Function to extract Arbeitspaket from the second column of the DataFrame
def get_arbeitspaket(df):
    arbeitspaket = []
    counter = 0
    for line in df.values:
        line_to_string = str(line[1])
        id_string = str(line[0])

        if line_to_string != 'nan' and id_string != 'nan':
            try:
                # Attempt to convert to float
                id_float = float(id_string)
                # Check if the number is an integer by comparing it to its int version
                if not id_float.is_integer():
                    arbeitspaket.append(line_to_string)
            except ValueError:
                # If it cannot be converted to float, check if it is a hierarchical number
                if '.' in id_string:
                    arbeitspaket.append(line_to_string)
        else:
            if counter == 0 or "Summe" in line_to_string:
                arbeitspaket.append(line_to_string)
                counter += 1

    return arbeitspaket


def get_workers_pre_defined2(df):
    pre_define_workers = []
    id_com = get_name(df, "Aufteilung")
    for line in df.values:
        line_to_string = str(line[id_com])
        workers_split = line_to_string.split(";")
        pre_de_temp = []

        if str(line[0]) == "nan":
            continue

        # Check if line[0] is a valid non-integer index
        try:
            line_0_as_float = float(line[0])
            is_valid_index = not line_0_as_float.is_integer()  # Only valid if not a whole number
        except ValueError:
            is_valid_index = True  # If it can't be converted to a float, it's a valid string (e.g. "3.1.1")

        for w_s in workers_split:
            # Append if w_s is not "nan" or "Aufteilung" and line[0] is a valid non-integer index
            if w_s != "nan" and w_s != "Aufteilung" and is_valid_index:
                pre_de_temp.append(w_s)
            if w_s == "nan" and w_s != "Aufteilung" and is_valid_index:
                pre_de_temp.append("0")

        # If there are valid workers in this line, add them to the result
        if pre_de_temp:
            pre_define_workers.append(pre_de_temp)

    return pre_define_workers


def get_name(df, name):
    index = 0
    counter = 0
    for line in df.values.__array__():
        if counter < 1:
            counter += 1
            for ll in line:
                if ll == name:
                    return index
                index += 1
        else:
            print(f"Error, Company/University not Found: {name}")
            exit(1)


def get_all_names(df):
    global total_id
    lista = []
    try:
        for line in df.values.__array__():
            collecting = False
            for i, name in enumerate(line):
                if i == 2:  # Start collecting from the second position (index 2)
                    collecting = True

                if collecting:
                    if isinstance(name, str) and name.lower() == "total":
                        total_id = i
                        return lista  # Stop collecting when "total" is encountered

                    if pd.notna(name) and isinstance(name, str) and not name.isdigit():
                        lista.append(name)

    except AttributeError as e:
        print(f"Error accessing DataFrame values: {e}")
        return []


def get_Company_hours(df, com_id):
    Company_worker = []
    Company_hour = []
    index = 0
    for line in df.values.__array__():
        line_to_string = str(line[com_id])

        if type(line[1]) == str and "Summe der Personenmonate" in line[1]:
            break

        if line_to_string != 'nan' and str(line[0]) != 'nan':
            infos = line[com_id].split(" ")
            Company_worker.append(infos[0])
            Company_hour.append(int(infos[1][2]))

    return Company_hour


def get_chosen_worker(id):
    for w in worker.list_of_workers:
        if w.id == id:
            return w


def get_Company_hours_and_worker(df, com_id):
    Company_worker = []
    Company_hour = []
    ids = []
    Nrs = []
    index = 0
    for line in df.values.__array__():
        line_to_string = str(line[com_id])

        if type(line[1]) == str and "Summe der Personenmonate" in line[1]:
            break

        if line_to_string != 'nan' and str(line[0]) != 'nan':
            infos = line[com_id].split(" ")
            chosen_worker = get_chosen_worker(int(infos[0][2]))
            Company_worker.append(chosen_worker)
            Company_hour.append(int(infos[1][2]))
            ids.append(line[0])
            Nrs.append(line[1])

    return Company_worker, Company_hour, ids, Nrs


def get_workers_pre_defined(df):
    pre_define_workers = []
    id_com = get_name(df, "Aufteilung")
    for line in df.values.__array__():
        line_to_string = str(line[id_com])
        workers_split = line_to_string.split(";")
        pre_de_temp = []
        for w_s in workers_split:
            if w_s != "nan" and w_s != "Aufteilung" and str(line[0]) != 'nan':
                pre_de_temp.append(w_s)
                continue
            if str(line[0]) != 'nan' and not float(line[0]).is_integer():
                pre_de_temp.append(0)
        if len(pre_de_temp) != 0:
            pre_define_workers.append(pre_de_temp)
    return pre_define_workers


# Function to extract APES values from the fifth column of the DataFrame
def get_Company(df, name):
    Company = []
    com_id = get_name(df, name)
    counter = 0

    index = 0
    for line in df.values.__array__():
        line_to_string = str(line[com_id])

        if type(line[1]) == str and "Summe der Personenmonate" in line[1]:
            break

        if line_to_string != 'nan' and str(line[0]) != 'nan':
            try:
                id_float = float(line[0])
                if not id_float.is_integer():
                    Company.append(line_to_string)
            except ValueError:
                if '.' in line[0]:
                    Company.append(line_to_string)
        else:
            try:
                if line_to_string == 'nan' and str(line[0]) != 'nan' and not float(line[0]).is_integer():
                    Company.append(0)
            except ValueError:
                if '.' in line[0]:
                    Company.append(0)
    return Company


# Define a basic color map for indexed colors if needed
INDEXED_COLOR_MAP = {
    8: 'Black',
    13: 'Red',
    14: 'Blue',
    15: 'Yellow',
    # Add more indexed colors if needed
}


def rgb_to_color_name(rgb):
    # Ensure `rgb` is in a format like "RRGGBB"
    if not rgb:
        return 'No Color'

    if rgb.startswith('FF'):  # Handle cases where RGB is in format 'FFRRGGBB'
        rgb = rgb[2:]

    try:
        hex_color = f"#{rgb[:6]}"  # Get the first 6 characters, ignoring the alpha channel if present
        return webcolors.hex_to_name(hex_color)
    except ValueError:
        # If the exact color name is not found, get the closest match
        closest_name = None
        min_distance = float('inf')
        rgb_tuple = webcolors.hex_to_rgb(hex_color)
        for color_name, color_hex in webcolors.CSS3_NAMES_TO_HEX.items():
            distance = sum((c1 - c2) ** 2 for c1, c2 in zip(rgb_tuple, webcolors.hex_to_rgb(color_hex)))
            if distance < min_distance:
                closest_name = color_name
                min_distance = distance
        return closest_name


def get_color_of_company(df, filepath, name):
    # Load the workbook and select the only sheet
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active  # Since there's only one sheet, we can use .active
    com_id = get_name(df, name) + 1
    Company_worker = []
    Company_hour = []
    colors = []

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):  # Assuming header in the first row
        cell = row[com_id - 1]  # Adjust for 0-based indexing

        if isinstance(cell.value, str) and "Summe der Personenmonate" in cell.value:
            break

        if cell.value is not None:
            Company_worker.append(cell.value)
            color = cell.font.color
            color_name = 'No Color'

    return Company_worker, Company_hour, colors


# Function to extract date information from the Excel file
def get_dates(filepath):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(filepath, data_only=True)
    fs = wb.active
    fs_count_row = fs.max_row
    fs_count_col = fs.max_column
    indexs_begin = []
    indexs_end = []

    # Iterate through rows and columns to find date blocks
    for row in range(6, fs_count_row + 1):
        end = 0
        start = 0
        for column in range(total_id + 1, fs_count_col + 1):
            cell_color = fs.cell(column=column, row=row)
            # Check cell color to identify date blocks
            if cell_color.fill.start_color.indexed == 8:
                if column == 1:
                    break
                if start == 0:
                    start = column
                    indexs_begin.append(column)
                    end = column
                end += 1
            else:
                if end > 0:  # If the block ends, record its end index
                    indexs_end.append(end)
                    break
                else:
                    continue

    return [indexs_begin, indexs_end]


# Function to filter out strings containing "jahr" from a list of strings
def filter_strings(string_list):
    new_years_array = []
    for y in string_list:
        if str(y) != 'nan':
            new_years_array.append(y)
    return new_years_array


# Function to map month abbreviations to numerical representation
def get_month_num(month):
    month_mapping = {
        "Jan": "01",
        "Feb": "02",
        "Mrz": "03",
        "Apr": "04",
        "Mai": "05",
        "Jun": "06",
        "Jul": "07",
        "Aug": "08",
        "Sep": "09",
        "Okt": "10",
        "Nov": "11",
        "Dez": "12",
        "Fev": "02",
        "Mar": "03",
        "Abr": "04",
        "Ago": "08",
        "Set": "09",
        "Out": "10",
    }
    return month_mapping.get(month, None)


def count_months_per_year(months_list):
    # Initialize a list to store the count of months per year
    months_per_year = []

    # Initialize variables to track the current year and month count
    current_year = 1
    current_month_count = 0
    counter = 0
    first_month = months_list[0]
    new_year_month = months_list[1]

    first = 0

    # Iterate through the months list
    for month in months_list:

        if first == 0:
            current_month_count += 1
            first = 1

        counter += 1
        # Check if the current month is December or if it's the last month in the list
        if month == 'Dez' or counter == len(months_list):
            # Append the month count for the current year to the list
            months_per_year.append(current_month_count)

            # Reset the month count for the next year
            current_month_count = 0

            # Increment the current year
            current_year += 1

        # Increment the month count for the current year
        current_month_count += 1

    return months_per_year


def get_years(df):
    years = df.values.__array__()[1]
    return years


# Function to convert dates to Unix format
def get_dates_unix(df, lista):
    list_begin = np.array(lista[0])
    list_end = np.array(lista[1])
    diff = list_begin[0]
    diff2 = list_end[0] + diff
    list_begin = (list_begin - diff)  # 6 see table
    list_end = (list_end - (diff + 1))  # 7 see table
    lista_anos_begin = []
    lista_anos_end = []
    months = np.array(df.values.__array__()[2])
    years = get_years(df)
    years = filter_strings(years)
    lista_begin_datas_unix = []
    lista_end_datas_unix = []

    temp_array = []
    for m in months:
        if str(m) != 'nan':
            temp_array.append(m)
    months = temp_array

    months_per_year_count = count_months_per_year(months)

    # Extract year information from column labels
    for m in list_begin:
        months_passed = months_per_year_count[0]
        index = 0
        while index < len(months_per_year_count):
            if m < months_passed:
                lista_anos_begin.append(index)
                break
            index += 1
            months_passed += months_per_year_count[index]

    for m in list_end:
        months_passed = months_per_year_count[0]
        index = 0
        while index <= len(months_per_year_count):
            if m < months_passed:
                lista_anos_end.append(index)
                break
            index += 1
            months_passed += months_per_year_count[index]

    end_day_mapping = {
        "Jan": "31", "Feb": "28", "Mrz": "31", "Apr": "30", "Mai": "31", "Jun": "30",
        "Jul": "31", "Aug": "31", "Sep": "30", "Okt": "31", "Nov": "30", "Dez": "31"
    }

    # Convert date components to Unix format
    for index in range(len(list_begin)):
        data = str("01." + get_month_num(months[list_begin[index]]) + "." + str(years[lista_anos_begin[index]]))
        lista_begin_datas_unix.append(data)

    for index in range(len(list_end)):
        day = end_day_mapping.get(months[list_end[index]])
        month = get_month_num(months[list_end[index]])
        year = years[lista_anos_end[index]]
        if int(year) % 4 == 0 and month == "Feb":
            data = str(day) + "." + "29" + "." + str(year)
            lista_end_datas_unix.append(data)
        else:
            data = str(day) + "." + str(month) + "." + str(year)
            lista_end_datas_unix.append(data)

    return [lista_begin_datas_unix, lista_end_datas_unix], months_per_year_count, lista_anos_begin, lista_anos_end
